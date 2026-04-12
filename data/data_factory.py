import json
from pathlib import Path
from faker import Faker
from config.config import config

fake = Faker()


class DataFactory:

    DATA_DIR = config.DATA_DIR

    @classmethod
    def get_json(cls, fixture_name: str) -> dict:
        path = cls.DATA_DIR / "json" / f"{fixture_name}.json"
        if not path.exists():
            raise FileNotFoundError(f"Fixture not found: {path}")
        with open(path) as f:
            return json.load(f)

    @classmethod
    def generate_user(cls, role: str = "customer") -> dict:
        return {
            "first_name": fake.first_name(),
            "last_name":  fake.last_name(),
            "email":      fake.email(),
            "password":   fake.password(length=12),
            "phone":      fake.phone_number(),
            "role":       role,
        }

    @classmethod
    def generate_product(cls) -> dict:
        return {
            "name":        fake.word().capitalize() + " Product",
            "price":       round(fake.pyfloat(
                               min_value=1,
                               max_value=1000,
                               right_digits=2), 2),
            "description": fake.sentence(),
            "category":    fake.word(),
        }

    @classmethod
    def generate_order(cls, user_id: int = 1,
                        product_id: int = 1) -> dict:
        return {
            "user_id":    user_id,
            "product_id": product_id,
            "quantity":   fake.random_int(min=1, max=10),
            "status":     "pending",
        }

    @classmethod
    def generate_address(cls) -> dict:
        return {
            "street":  fake.street_address(),
            "city":    fake.city(),
            "state":   fake.state(),
            "country": fake.country(),
            "zipcode": fake.zipcode(),
        }

    @classmethod
    def generate_credit_card(cls) -> dict:
        return {
            "number":   fake.credit_card_number(),
            "expiry":   fake.credit_card_expire(),
            "cvv":      fake.credit_card_security_code(),
            "provider": fake.credit_card_provider(),
        }

    @classmethod
    def get_excel(cls, filename: str,
                   sheet: str = "Sheet1") -> list:
        """Load Excel test data as list of dicts."""
        import openpyxl
        path = cls.DATA_DIR / "excel" / filename
        wb   = openpyxl.load_workbook(path)
        ws   = wb[sheet] if sheet in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = rows[0]
        return [
            dict(zip(headers, row))
            for row in rows[1:]
            if any(cell is not None for cell in row)
        ]